"""
generate_profile_js_from_xlsx_v2.py

Reads HD2_Profile_Tuner_Condensed_v2.xlsx-style workbooks and outputs a FULL RESOLVED JS profile:
export default { name, rules, macro, micro };

Sheets:
- META: ProfileName (B1), UseRulesFromBase (B2)
- MACRO: group | key | weight
- MICRO: kind(tag|id) | key | weight
- SCOPED_EQUIP: pool | categoryKey | kind(tag|id) | key | weight
- SCOPED_STRAT: categoryKey | kind(tag|id) | key | weight

Rules: currently emitted as a sane default block (same idea as previous script). You can add a RULES sheet later.
"""

from __future__ import annotations
import sys
from openpyxl import load_workbook

def fnum(v, default=1.0):
    if v is None or str(v).strip() == "":
        return float(default)
    return float(v)

def read_macro(ws):
    out = {
        "primarySubcategory": {},
        "secondarySubcategory": {},
        "throwableSubcategory": {},
        "stratagemTop": {},
        "offensiveSub": {},
        "defensiveSub": {},
        "supplySub": {},
    }
    for r in range(2, ws.max_row + 1):
        group = ws.cell(r, 1).value
        key = ws.cell(r, 2).value
        w = ws.cell(r, 3).value
        if not group or not key:
            continue
        group = str(group).strip()
        key = str(key).strip()
        if group not in out:
            # allow extra groups without crashing; ignore
            continue
        out[group][key] = fnum(w, 1.0)
    return out

def read_micro(ws):
    tag = {}
    idw = {}
    for r in range(2, ws.max_row + 1):
        kind = ws.cell(r, 1).value
        key = ws.cell(r, 2).value
        w = ws.cell(r, 3).value
        if not kind or not key:
            continue
        kind = str(kind).strip().lower()
        key = str(key).strip()
        if kind == "tag":
            tag[key] = fnum(w, 1.0)
        elif kind == "id":
            idw[key] = fnum(w, 1.0)
    return tag, idw

def read_scoped_equip(ws):
    pools = { "primaries": {}, "secondaries": {}, "throwables": {} }
    for r in range(2, ws.max_row + 1):
        pool = ws.cell(r, 1).value
        cat = ws.cell(r, 2).value
        kind = ws.cell(r, 3).value
        key = ws.cell(r, 4).value
        w = ws.cell(r, 5).value
        if not pool or not cat or not kind or not key:
            continue
        pool = str(pool).strip().lower()
        if pool not in pools:
            continue
        cat = str(cat).strip()
        kind = str(kind).strip().lower()
        key = str(key).strip()
        node = pools[pool].setdefault(cat, {"tag": {}, "id": {}})
        if kind == "tag":
            node["tag"][key] = fnum(w, 1.0)
        elif kind == "id":
            node["id"][key] = fnum(w, 1.0)
    return pools

def read_scoped_strat(ws):
    cats = {}
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        kind = ws.cell(r, 2).value
        key = ws.cell(r, 3).value
        w = ws.cell(r, 4).value
        if not cat or not kind or not key:
            continue
        cat = str(cat).strip()
        kind = str(kind).strip().lower()
        key = str(key).strip()
        node = cats.setdefault(cat, {"tag": {}, "id": {}})
        if kind == "tag":
            node["tag"][key] = fnum(w, 1.0)
        elif kind == "id":
            node["id"][key] = fnum(w, 1.0)
    return cats

def js_obj(d, indent=2, level=0):
    sp = " " * (indent * level)
    if isinstance(d, dict):
        if not d:
            return "{}"
        lines = ["{"]
        for k,v in d.items():
            key = f'"{k}"' if ("/" in k or " " in k or "-" in k) else (k if k.isidentifier() else f'"{k}"')
            lines.append(f'{sp}{" " * indent}{key}: {js_obj(v, indent, level+1)},')
        lines.append(f'{sp}}}')
        return "\n".join(lines)
    if isinstance(d, str):
        return f'"{d}"'
    if isinstance(d, bool):
        return "true" if d else "false"
    if isinstance(d, (int, float)):
        return f"{float(d):.2f}"
    if d is None:
        return "null"
    raise TypeError(type(d))

def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_profile_js_from_xlsx_v2.py input.xlsx output.profile.js")
        sys.exit(2)

    in_xlsx = sys.argv[1]
    out_js = sys.argv[2]

    wb = load_workbook(in_xlsx, data_only=True)

    meta = wb["META"]
    name = meta["B1"].value or "Custom"
    use_rules_from_base = str(meta["B2"].value or "true").strip().lower() == "true"

    macro = read_macro(wb["MACRO"])
    tag, idw = read_micro(wb["MICRO"])
    equip_scoped = read_scoped_equip(wb["SCOPED_EQUIP"])
    strat_scoped = read_scoped_strat(wb["SCOPED_STRAT"])

    micro = {
        "tag": tag,
        "id": idw,
        "pools": {
            "primaries": {"tag": {}, "id": {}, "categories": equip_scoped["primaries"]},
            "secondaries": {"tag": {}, "id": {}, "categories": equip_scoped["secondaries"]},
            "throwables": {"tag": {}, "id": {}, "categories": equip_scoped["throwables"]},
            "stratagems": {"tag": {}, "id": {}, "categories": strat_scoped},
        }
    }

    rules = {
        "useEvenStratagemTop": True,
        "maxSupportWeapons": 1,
        "maxBackpackSlot": 1,
        "maxSupplyStratagems": 2,
        "preventDoubleExosuit": True,
        "soloSiloCannotBeOnlySupport": True,
        "enforceOneHandedIfRequired": True,
        "ALLOW_EXPENDABLE_AS_GUARANTEED_SUPPORT": False,
    } if use_rules_from_base else {"useEvenStratagemTop": True}

    profile = {"name": str(name), "rules": rules, "macro": macro, "micro": micro}
    js = "export default " + js_obj(profile, indent=2, level=0) + ";\n"
    with open(out_js, "w", encoding="utf-8") as f:
        f.write(js)
    print(f"Wrote: {out_js}")

if __name__ == "__main__":
    main()
