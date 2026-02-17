#!/usr/bin/env python3
"""make_audit_xlsx.py (v7)

Takes audit exports:
- hd2_audit_items_*.csv
- hd2_audit_tags_*.csv (optional)

Outputs an XLSX workbook with:
- Tables for Items (All + by kind)
- Summary tables by category/subcategory
- Charts sheet with readable pies (legend-only) PLUS percent tables under each chart
- Tag visuals (bar + pie) + percent tables

Notes:
- Excel does not support percentages inside legend text. The percent tables below charts act as an "extended legend".
- Stratagem Top Type pie uses deterministic order: Supply (blue), Defensive (green), Offensive (red)
  (Excel assigns colors by slice order; openpyxl cannot reliably hard-set slice colors.)
"""

import argparse
import csv
from collections import defaultdict, Counter

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import PieChart, BarChart, Reference


def read_csv(path):
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        raise ValueError(f"Empty CSV: {path}")
    return rows[0], rows[1:]


def add_sheet_with_table(wb, title, header, rows, table_name):
    ws = wb.create_sheet(title)
    ws.append(header)
    for r in rows:
        ws.append(r)

    for col_idx, h in enumerate(header, 1):
        if col_idx <= 26:
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = max(12, min(46, len(h) + 2))

    end_row = len(rows) + 1
    end_col = len(header)
    ref = f"A1:{chr(64 + end_col)}{end_row}"

    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    return ws


def add_summary_sheet(wb, title, rows, table_name):
    header = ["top", "sub", "total_count", "sum_odds_percent"]
    return add_sheet_with_table(wb, title, header, rows, table_name)


def build_group_summary(items_rows, idx, kind_value, group_by_top=True):
    grouped = defaultdict(lambda: {"count": 0, "odds_sum": 0.0})
    for r in items_rows:
        if r[idx["kind"]] != kind_value:
            continue
        top = r[idx["top"]] if group_by_top else ""
        sub = r[idx["sub"]]
        try:
            c = int(float(r[idx["count"]] or 0))
        except Exception:
            c = 0
        try:
            o = float(r[idx["odds_percent"]] or 0.0)
        except Exception:
            o = 0.0
        grouped[(top, sub)]["count"] += c
        grouped[(top, sub)]["odds_sum"] += o

    out = []
    for (top, sub), v in grouped.items():
        out.append([top, sub, v["count"], round(v["odds_sum"], 6)])
    out.sort(key=lambda x: (x[0], x[1]))
    return out


def write_chart_table(ws, start_row, title, labels, values):
    # Returns (data_start_row, data_end_row, total_row)
    ws.cell(row=start_row, column=1, value=title)
    ws.cell(row=start_row + 1, column=1, value="Label")
    ws.cell(row=start_row + 1, column=2, value="Value")
    ws.cell(row=start_row + 1, column=3, value="Percent")
    r = start_row + 2
    total = sum(float(v) for v in values) if values else 0.0
    for lab, val in zip(labels, values):
        v = float(val)
        pct = (v / total * 100.0) if total > 0 else 0.0
        ws.cell(row=r, column=1, value=str(lab))
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=3, value=round(pct, 4))
        r += 1

    # total row
    ws.cell(row=r, column=1, value="Total")
    ws.cell(row=r, column=2, value=total)
    ws.cell(row=r, column=3, value=100.0 if total > 0 else 0.0)

    # Format percent column as %
    for rr in range(start_row + 2, r + 1):
        ws.cell(row=rr, column=3).number_format = '0.00"%"'

    return start_row + 2, r - 1, r


def add_pie(ws, title, data_start_row, data_end_row, anchor):
    pie = PieChart()
    pie.title = title
    data = Reference(ws, min_col=2, min_row=data_start_row - 1, max_row=data_end_row)
    labels = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = None  # legend-only
    ws.add_chart(pie, anchor)


def add_bar(ws, title, data_start_row, data_end_row, anchor):
    bar = BarChart()
    bar.type = "col"
    bar.title = title
    bar.y_axis.title = "Value"
    data = Reference(ws, min_col=2, min_row=data_start_row - 1, max_row=data_end_row)
    cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, anchor)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--items", required=True, help="Item odds CSV from audit tool")
    ap.add_argument("--tags", required=False, help="Tag odds CSV from audit tool")
    ap.add_argument("--slot1", required=False, help="Slot 1 stratagem picks CSV from audit tool")
    ap.add_argument("--out", required=True, help="Output .xlsx path")
    args = ap.parse_args()

    items_header, items_rows = read_csv(args.items)
    idx = {h: i for i, h in enumerate(items_header)}
    for need in ["kind", "top", "sub", "id", "name", "count", "odds_percent"]:
        if need not in idx:
            raise ValueError(f"Items CSV missing column '{need}'. Export items CSV from audit tool v3+." )

    by_kind = defaultdict(list)
    for r in items_rows:
        by_kind[r[idx["kind"]]].append(r)

    def sort_rows(rows):
        def key(r):
            try:
                return float(r[idx["odds_percent"]])
            except Exception:
                return 0.0
        return sorted(rows, key=key, reverse=True)

    wb = Workbook()
    wb.remove(wb.active)

    add_sheet_with_table(wb, "Items_All", items_header, items_rows, "ItemsAll")
    for kind, title in [("Primary", "Primaries"), ("Secondary", "Secondaries"), ("Throwable", "Throwables"), ("Stratagem", "Stratagems")]:
        add_sheet_with_table(wb, title, items_header, sort_rows(by_kind.get(kind, [])), title + "Tbl")

    prim_sum = build_group_summary(items_rows, idx, "Primary", group_by_top=False)
    sec_sum = build_group_summary(items_rows, idx, "Secondary", group_by_top=False)
    thr_sum = build_group_summary(items_rows, idx, "Throwable", group_by_top=False)
    for r in prim_sum: r[0] = "Primary"
    for r in sec_sum: r[0] = "Secondary"
    for r in thr_sum: r[0] = "Throwable"
    add_summary_sheet(wb, "Primaries_BySubcategory", prim_sum, "PrimBySub")
    add_summary_sheet(wb, "Secondaries_BySubcategory", sec_sum, "SecBySub")
    add_summary_sheet(wb, "Throwables_BySubcategory", thr_sum, "ThrBySub")

    strat_sum = build_group_summary(items_rows, idx, "Stratagem", group_by_top=True)
    add_summary_sheet(wb, "Stratagems_ByTopSub", strat_sum, "StratByTopSub")

    tag_rows = None
    tag_idx = None
    if args.tags:
        tags_header, tags_rows = read_csv(args.tags)
        add_sheet_with_table(wb, "Tags", tags_header, tags_rows, "TagsTbl")
        tag_rows = tags_rows
        tag_idx = {h: i for i, h in enumerate(tags_header)}
    slot1_rows = None
    slot1_idx = None
    if args.slot1:
        slot1_header, slot1_rows = read_csv(args.slot1)
        add_sheet_with_table(wb, "Stratagems_Slot1", slot1_header, slot1_rows, "StratSlot1Tbl")
        slot1_idx = {h: i for i, h in enumerate(slot1_header)}


    # Charts
    ws = wb.create_sheet("Charts")

    # Primaries pie
    prim_counts = [(r[1], r[2]) for r in prim_sum]  # (sub, total_count)
    prim_counts.sort(key=lambda x: x[1], reverse=True)
    p_labels = [x[0] for x in prim_counts]
    p_values = [x[1] for x in prim_counts]
    ds, de, total_row = write_chart_table(ws, 1, "Primaries by Subcategory (count)", p_labels, p_values)
    add_pie(ws, "Primaries by Subcategory", ds, de, "D2")

    # Stratagem top pie (Supply blue, Defensive green, Offensive red)
    strat_top_counter = Counter()
    for top, sub, total_count, _ in strat_sum:
        strat_top_counter[top] += int(total_count)

    ORDER = ["Supply", "Defensive", "Offensive"]
    st_labels = [k for k in ORDER if k in strat_top_counter]
    st_labels += [k for k in strat_top_counter.keys() if k not in st_labels]
    st_values = [strat_top_counter[k] for k in st_labels]

    ds2, de2, total_row2 = write_chart_table(ws, total_row + 3, "Stratagems by Top Type (count)", st_labels, st_values)
    add_pie(ws, "Stratagems by Top Type", ds2, de2, "D18")

    # Stratagem subtype pie (Top 10 + Other)
    strat_sub_counter = Counter()
    for top, sub, total_count, _ in strat_sum:
        strat_sub_counter[f"{top} / {sub}"] += int(total_count)
    most = strat_sub_counter.most_common(10)
    other = sum(strat_sub_counter.values()) - sum(v for _, v in most)
    ss_labels = [k for k, _ in most] + (["Other"] if other > 0 else [])
    ss_values = [v for _, v in most] + ([other] if other > 0 else [])
    ds3, de3, total_row3 = write_chart_table(ws, total_row2 + 3, "Stratagems by Subtype (Top 10 + Other, count)", ss_labels, ss_values)
    add_pie(ws, "Stratagems by Subtype (Top 10)", ds3, de3, "D34")

    # Tags visuals
    if tag_rows and tag_idx and "tag" in tag_idx:
        tag_col = tag_idx.get("tag", 0)
        avg_col = tag_idx.get("avg_per_run", None)

        parsed = []
        for r in tag_rows:
            tag = r[tag_col]
            try:
                avg = float(r[avg_col]) if avg_col is not None else float(r[2])
            except Exception:
                continue
            parsed.append((tag, avg))

        parsed.sort(key=lambda x: x[1], reverse=True)

        # Bar: Top 25
        top25 = parsed[:25]
        t25_labels = [t for t, _ in top25]
        t25_values = [v for _, v in top25]
        ds4, de4, total_row4 = write_chart_table(ws, total_row3 + 3, "Top 25 Tags (avg per run)", t25_labels, t25_values)
        add_bar(ws, "Top 25 Tags (avg per run)", ds4, de4, "D50")

        # Pie: Top 10 + Other
        top10 = parsed[:10]
        other_val = sum(v for _, v in parsed[10:])
        tp_labels = [t for t, _ in top10] + (["Other"] if other_val > 0 else [])
        tp_values = [v for _, v in top10] + ([other_val] if other_val > 0 else [])
        ds5, de5, total_row5 = write_chart_table(ws, total_row4 + 3, "Tags (Top 10 + Other, avg per run)", tp_labels, tp_values)
        add_pie(ws, "Tags (Top 10 + Other)", ds5, de5, "D68")


    # Slot 1 Stratagem Top-Type (if provided)
    if args.slot1 and slot1_rows and slot1_idx and "top" in slot1_idx and "count" in slot1_idx:
        ws = wb["Charts"]
        last_row = ws.max_row + 3
        ctop = Counter()
        for r in slot1_rows:
            try:
                top = r[slot1_idx["top"]]
                cnt = float(r[slot1_idx["count"]])
            except Exception:
                continue
            ctop[top] += cnt

        ORDER = ["Supply","Defensive","Offensive"]
        labels = [k for k in ORDER if k in ctop]
        labels += [k for k in ctop.keys() if k not in labels]
        values = [ctop[k] for k in labels]
        ds6, de6, total_row6 = write_chart_table(ws, last_row, "Slot 1 Stratagem Top Type (count)", labels, values)
        add_pie(ws, "Slot 1 Stratagem Top Type", ds6, de6, "D84")

    wb.save(args.out)
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
